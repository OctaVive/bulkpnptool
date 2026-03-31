from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

from flask import Flask, Response, render_template, request, send_from_directory

from pe_processor import get_pe_processor


app = Flask(__name__)


@dataclass
class ParsedNumber:
    raw_input: str
    is_wildcard: bool
    lookup_national: Optional[str]  # national format with leading 0, digits only
    export_value: str               # exact value to write into CSV


SPECIAL_LOCATION_PREFIXES: Tuple[str, ...] = (
    "050",
    "0521",
    "0522",
    "0524",
    "0527",
    "0598",
    "0599",
)


def _clean_input_line(line: str) -> str:
    # Strip surrounding whitespace but otherwise keep as-is for wildcard export.
    return line.strip()


def _normalize_for_lookup(raw: str) -> Tuple[Optional[str], bool]:
    """
    Normalize a user-entered phone number for PE lookup.

    Behaviour:
    - Remove all characters except digits and 'X'/'x'.
    - Convert '+31' prefix to '0'.
    - Convert 9-digit numbers (without leading 0) to 10-digit by prefixing '0'.
    - For wildcard blocks, 'X' is only used for lookup by treating it as '0'.

    Returns:
        (national_number_with_leading_zero_or_None, is_wildcard)
    """
    if not raw:
        return None, False

    s = raw.strip()

    # Handle +31 country code (only for normalization, NOT export).
    if s.startswith("+31"):
        s = "0" + s[3:]

    # Keep digits and X/x only.
    cleaned_chars: List[str] = []
    is_wildcard = False
    for ch in s:
        if ch.isdigit():
            cleaned_chars.append(ch)
        elif ch in {"X", "x"}:
            cleaned_chars.append("X")
            is_wildcard = True
        # all other characters are discarded for lookup purposes

    if not cleaned_chars:
        return None, False

    cleaned = "".join(cleaned_chars)

    # Handle 0031 / 31 country code variants that were not covered by the +31
    # logic above. This ensures inputs like 3144... are treated as 044...
    # for lookup, so the correct PE region (e.g. 044 -> PE60) is chosen.
    if cleaned.startswith("0031"):
        cleaned = "0" + cleaned[4:]
    elif cleaned.startswith("31"):
        cleaned = "0" + cleaned[2:]

    # For wildcard lookup, treat X as 0.
    lookup_str = cleaned.replace("X", "0")

    # Ensure we have a national-format number (leading 0, 10 digits) for lookup.
    digits_only = "".join(ch for ch in lookup_str if ch.isdigit())

    # If we see a local 050-range number (starting with "50"), make it explicit
    # national format "050..." so that it is treated as ambiguous and triggers
    # the 050 / special-prefix selection flow. We do not apply the generic
    # 9-digit rule again in this case to avoid ending up with "0050...".
    if digits_only.startswith("50") and not digits_only.startswith("050"):
        digits_only = "0" + digits_only
    elif len(digits_only) == 9:
        # Local form, add leading 0.
        digits_only = "0" + digits_only
    # If already 10 digits starting with 0, leave as-is; otherwise we still try
    # prefix resolution with whatever we have, as long as length >= 3.

    if len(digits_only) < 3:
        return None, is_wildcard

    return digits_only, is_wildcard


def _parse_numbers(raw_numbers: str) -> List[ParsedNumber]:
    """
    Parse and normalize raw user input into structures used for lookup and export.
    """
    parsed: List[ParsedNumber] = []

    for line in raw_numbers.splitlines():
        cleaned_line = _clean_input_line(line)
        if not cleaned_line:
            continue

        lookup_national, is_wildcard = _normalize_for_lookup(cleaned_line)

        export_value = cleaned_line

        # Special-location prefixes (050 / 0521 / 0522 / 0524 / 0527 / 0598 / 0599)
        # are handled first to preserve the existing 050 behaviour and avoid
        # changing ambiguous prefix handling.
        if lookup_national and lookup_national.startswith(tuple(SPECIAL_LOCATION_PREFIXES)):
            if is_wildcard:
                # For 050 wildcard blocks (e.g. 506882XX) keep the user's text
                # exactly as entered; only PE resolution uses the normalized
                # 050 form.
                export_value = cleaned_line
            else:
                digits = "".join(ch for ch in lookup_national if ch.isdigit())
                if digits:
                    # For special-location prefixes we keep exporting the
                    # "local" 9-digit form (strip the first digit when a full
                    # 10-digit national number is present), matching the
                    # original 050 behaviour.
                    if len(digits) >= 10:
                        export_value = digits[1:10]
                    elif len(digits) == 9:
                        export_value = digits[1:]
                    else:
                        export_value = digits
        else:
            # General rule for all non-050 numbers (wildcard or not):
            # normalize any format to a 9-character local representation by
            # taking the last 9 characters of the cleaned number (digits and X).
            export_clean = "".join(
                ch for ch in cleaned_line if ch.isdigit() or ch in {"X", "x"}
            )
            if len(export_clean) >= 9:
                export_value = export_clean[-9:]
            elif len(export_clean) > 0:
                export_value = export_clean

        parsed.append(
            ParsedNumber(
                raw_input=cleaned_line,
                is_wildcard=is_wildcard,
                lookup_national=lookup_national,
                export_value=export_value,
            )
        )

    return parsed


def _normalize_for_block_export(raw: str) -> Optional[str]:
    """
    Normalize a user-entered number/block to national display format for XLSX.

    Output keeps wildcard markers and aims for 10-character Dutch national
    format, e.g. 455688200 -> 0455688200 and 45568820X -> 045568820x.
    """
    if not raw:
        return None

    s = raw.strip()
    if not s:
        return None

    if s.startswith("+31"):
        s = "0" + s[3:]

    cleaned_chars: List[str] = []
    for ch in s:
        if ch.isdigit():
            cleaned_chars.append(ch)
        elif ch in {"X", "x"}:
            cleaned_chars.append("x")

    if not cleaned_chars:
        return None

    cleaned = "".join(cleaned_chars)

    if cleaned.startswith("0031"):
        cleaned = "0" + cleaned[4:]
    elif cleaned.startswith("31"):
        cleaned = "0" + cleaned[2:]

    if len(cleaned) == 9 and not cleaned.startswith("0"):
        cleaned = "0" + cleaned

    return cleaned


def _block_type_for_value(block_value: str) -> str:
    wildcard_count = block_value.count("x")
    if wildcard_count == 0:
        return "Single number"
    if wildcard_count == 1:
        return "10 numbers"
    if wildcard_count == 2:
        return "100 numbers"
    if wildcard_count == 3:
        return "1000 numbers"
    return f"{10 ** wildcard_count} numbers"


def _build_xlsx_blocks(raw_numbers: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Blocks"
    ws.append(["BlockType", "Block"])

    header_fill = PatternFill(fill_type="solid", start_color="E60000", end_color="E60000")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for line in raw_numbers.splitlines():
        normalized = _normalize_for_block_export(line)
        if not normalized:
            continue
        ws.append([_block_type_for_value(normalized), normalized])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _load_special_locations_by_prefix() -> Dict[str, List[Dict[str, str]]]:
    """
    Load special PE location lists per prefix (050/0521/0522/0524/0527/0598/0599).

    Returned format:
      {
        "050": [{"city": "...", "pe_code": "..."}, ...],
        "0521": [...],
        ...
      }
    """
    from pathlib import Path

    base_dir = Path(__file__).parent / "data"
    locations_by_prefix: Dict[str, List[Dict[str, str]]] = {}

    for prefix in SPECIAL_LOCATION_PREFIXES:
        locations: List[Dict[str, str]] = []
        locations_file = base_dir / f"{prefix}_locations.txt"
        if locations_file.exists():
            with locations_file.open("r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#"):
                        continue
                    parts = line.split(";")
                    if len(parts) < 2:
                        continue
                    city = parts[0].strip()
                    pe_code = parts[1].strip()
                    if not city or not pe_code:
                        continue
                    locations.append({"city": city, "pe_code": pe_code})

        # Sort alphabetically by city name so the dropdown is ordered.
        locations.sort(key=lambda loc: loc["city"].lower())
        locations_by_prefix[prefix] = locations

    return locations_by_prefix


def _collect_special_location_numbers(parsed_numbers: List[ParsedNumber]) -> List[ParsedNumber]:
    special_numbers: List[ParsedNumber] = []
    for p in parsed_numbers:
        if p.lookup_national and p.lookup_national.startswith(tuple(SPECIAL_LOCATION_PREFIXES)):
            special_numbers.append(p)
    return special_numbers


def _get_special_prefix(lookup_national: str) -> Optional[str]:
    """
    Determine which special prefix a normalized national lookup number belongs to.

    Order matters: check longer prefixes first (e.g. 0521 before 05x, etc.).
    """
    for prefix in sorted(SPECIAL_LOCATION_PREFIXES, key=len, reverse=True):
        if lookup_national.startswith(prefix):
            return prefix
    return None


def _build_csv_rows(
    parsed_numbers: List[ParsedNumber],
    operation: str,
    pbx_id_from: str,
    pbx_id_to: str,
    pe_050_overrides: Dict[str, str],
) -> List[List[str]]:
    pe_processor = get_pe_processor()

    rows: List[List[str]] = []
    row_number = 1

    op = operation.upper()

    for p in parsed_numbers:
        if not p.lookup_national:
            # Cannot resolve PE for this entry; skip it silently to avoid
            # generating potentially unsafe provisioning rows.
            continue

        # Handle ambiguous special-location numbers (050 / 0521 / 0522 / 0524 /
        # 0527 / 0598 / 0599) with user-provided PE override.
        if p.lookup_national.startswith(tuple(SPECIAL_LOCATION_PREFIXES)):
            pe_code = pe_050_overrides.get(p.lookup_national)
            if not pe_code:
                # No user selection; do not auto-resolve.
                continue
        else:
            pe_code = pe_processor.resolve_pe_code(p.lookup_national)
            if not pe_code:
                # No mapping; skip to avoid unsafe provisioning.
                continue

        # Determine which PBX ID(s) to use based on operation.
        if op == "ADD":
            logical_ops = [("addPNP", pbx_id_to)]
        elif op == "DELETE":
            logical_ops = [("deletePNP", pbx_id_from)]
        elif op == "MOVE":
            # MOVE → deletePNP + addPNP
            logical_ops = [
                ("deletePNP", pbx_id_from),
                ("addPNP", pbx_id_to),
            ]
        else:
            # Unknown operation; nothing to generate.
            continue

        for logical_op, pbx in logical_ops:
            rows.append(
                [
                    str(row_number),
                    logical_op,
                    p.export_value,
                    pe_code,
                    pbx,
                ]
            )
            row_number += 1

    return rows


@app.route("/assets/<path:filename>")
def asset(filename: str):
    """
    Serve static assets (e.g. Vodafone logo) from the local assets directory.
    """
    import os

    assets_dir = os.path.join(os.path.dirname(__file__), "assets")
    return send_from_directory(assets_dir, filename)


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "GET":
        return render_template("index.html")

    # POST: process numbers
    numbers_raw = request.form.get("numbers", "")
    operation = request.form.get("operation", "ADD")
    pbx_id_from = request.form.get("pbx_id_from", "")
    pbx_id_to = request.form.get("pbx_id_to", "")
    output_format = request.form.get("output_format", "csv").lower()

    if output_format == "xlsx":
        xlsx_data = _build_xlsx_blocks(numbers_raw)
        headers = {
            "Content-Disposition": 'attachment; filename="bulk_number_blocks.xlsx"',
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
        return Response(xlsx_data, headers=headers)

    parsed_numbers = _parse_numbers(numbers_raw)
    special_numbers = _collect_special_location_numbers(parsed_numbers)

    confirm_050 = request.form.get("confirm_050")

    if special_numbers and not confirm_050:
        # Render special-prefix selection popup.
        # Provide locations per prefix so each number row shows the correct
        # dropdown list (e.g. 050 -> 050_locations.txt only).
        locations_by_prefix = _load_special_locations_by_prefix()

        unique_prefixes = {
            _get_special_prefix(n.lookup_national)
            for n in special_numbers
            if n.lookup_national is not None
        }
        unique_prefixes.discard(None)

        # If mixed prefixes are present, hide the bulk location picker because
        # it can't safely apply across different *_locations.txt lists.
        bulk_locations: Optional[List[Dict[str, str]]] = None
        if len(unique_prefixes) == 1:
            only_prefix = next(iter(unique_prefixes), None)
            if only_prefix:
                bulk_locations = locations_by_prefix.get(only_prefix, [])
        # Use lookup_national as the key, which will be a 10-digit national-format
        # number beginning with 050, e.g. 0501234567. This matches the example
        # pe_0501234567=01.
        return render_template(
            "select_050.html",
            numbers_050=special_numbers,
            locations_by_prefix=locations_by_prefix,
            bulk_locations=bulk_locations,
            numbers_raw=numbers_raw,
            operation=operation,
            pbx_id_from=pbx_id_from,
            pbx_id_to=pbx_id_to,
        )

    # If we are here, either there were no special-prefix numbers or the user
    # already confirmed the PE codes for them.
    pe_050_overrides: Dict[str, str] = {}
    if confirm_050:
        # Collect all submitted PE codes for special-location numbers.
        for key, value in request.form.items():
            if not key.startswith("pe_"):
                continue
            national_number = key[len("pe_") :]
            pe_050_overrides[national_number] = value

    rows = _build_csv_rows(
        parsed_numbers=parsed_numbers,
        operation=operation,
        pbx_id_from=pbx_id_from,
        pbx_id_to=pbx_id_to,
        pe_050_overrides=pe_050_overrides,
    )

    # Generate CSV content.
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", lineterminator="\n")
    for row in rows:
        writer.writerow(row)

    csv_data = output.getvalue()
    output.close()

    # Serve as file download.
    headers = {
        "Content-Disposition": 'attachment; filename="bulk_pnp_operations.csv"',
        "Content-Type": "text/csv; charset=utf-8",
    }
    return Response(csv_data, headers=headers)


if __name__ == "__main__":
    # Simple built-in server for local testing and Docker.
    # Bind to 0.0.0.0 so the container's port 5000 is reachable.
    app.run(host="0.0.0.0", port=5000, debug=True)

